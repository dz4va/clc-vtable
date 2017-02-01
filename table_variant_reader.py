"""
# CLC Table variant data reader class uses given openpyxl
# sheet row to read data and return TableVariant object
# Author: George Dzavashvili
# Email: dzavashviligeorge@gmail.com
# Twitter: @redpix_
# How it works:
    TODO: Explain
"""
import config as cfg
import openpyxl as xl
from table_variant import TableVariant


class TableVariantReader(object):
    """Summary
    """

    def __init__(self, file_name):
        # Create workbook
        self.workbook = xl.load_workbook(file_name)
        # Set active sheet
        self.active_sheet = self.workbook.active

    def get_value(self, row, col):
        # print(type(self.active_sheet.cell(row=row, column=col).value))
        return self.active_sheet.cell(row=row, column=col).value

    def read_variant(self, row):
        columns = {
            cfg.V_REFPOS: self.get_value(row, 1),
            cfg.V_TYPE: self.get_value(row, 2),
            cfg.V_LENGTH: self.get_value(row, 3),
            cfg.V_REF: self.get_value(row, 4),
            cfg.V_ALLELE: self.get_value(row, 5),
            cfg.V_LINKAGE: self.get_value(row, 6),
            cfg.V_ZYGOSITY: self.get_value(row, 7),
            cfg.V_ALCOUNT: self.get_value(row, 8),
            cfg.V_COV: self.get_value(row, 9),
            cfg.V_FREQ: self.get_value(row, 10),
            cfg.V_FR_BAL: self.get_value(row, 11),
            cfg.V_AVG_QUAL: self.get_value(row, 12),
            cfg.V_ANOT: self.get_value(row, 13),
            cfg.V_CRC: self.get_value(row, 14),
            cfg.V_AAC: self.get_value(row, 15)
        }

        return TableVariant(**columns)
